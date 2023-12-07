import openpyxl
import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time

def getCurrentTime():
    # Get the current time and format it as a string
    return datetime.datetime.now().strftime("%Y/%m/%d %H:%M")



def get_trending():
    """
    coinpost.jpからトレンド記事を取得する。

    Returns:
    trending_links (list) :トレンドの記事リンクのリスト
    trending_titles (list):トレンド記事のタイトルのリスト。
    """

    # chromedriver settings
    options = Options()
    options.add_argument('--headless')

    # Set URL and create a Chrome webdriver
    url = "https://coinpost.jp/"
    driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
    driver.implicitly_wait(3)

    print("Launching browser...")

    # Access site
    driver.get(url)
    print("Accessing site...")
    time.sleep(3)

    # Get trending links and titles
    trending_links = []
    trending_titles = []
    for i in range(1, 6):
        link = driver.find_element(By.XPATH, f"/html/body/div[1]/div[4]/div/div[1]/div[2]/div/div/div/div[{i}]/a").get_attribute("href")
        title = driver.find_element(By.XPATH, f"/html/body/div[1]/div[4]/div/div[1]/div[2]/div/div/div/div[{i}]/a/span[2]").text
        trending_links.append(link)
        trending_titles.append(title)

    time.sleep(3)
    driver.quit()
    print("Browser closed.")

    return trending_links,trending_titles



def insert_trend_articles(article_links,article_titles):

    # Set the file path for the Excel file
    file_path = "トレンドニュース一覧.xlsx"

    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Find the next empty row in the worksheet
    row_num = 1
    while worksheet.cell(row=row_num, column=2).value is not None:
        row_num += 1

    # Write the current time to the first column of the next empty row
    current_time = getCurrentTime()
    worksheet["A" + str(row_num)] = current_time

    # Write the article titles and links to the worksheet
    for i in range(len(article_links)):
        row = row_num + i
        worksheet.cell(row=row, column=2).value = article_titles[i]
        worksheet.cell(row=row, column=2).hyperlink = article_links[i]

    # Save the changes to the Excel file
    workbook.save(file_path)
    print("Save the changes to the Excel file")

def main():
    article_links,article_titles = get_trending()
    insert_trend_articles(article_links,article_titles)

if __name__ == '__main__':
    main()
